from io import TextIOWrapper

import csv
import gzip
import hashlib
import json
import logging
import os
import tempfile
import openpyxl as xl

from seqr.utils.file_utils import mv_file_to_gs, file_iter
from seqr.views.utils.json_utils import create_json_response
from seqr.views.utils.permissions_utils import login_and_policies_required
from seqr.views.utils.terra_api_utils import anvil_enabled

logger = logging.getLogger(__name__)

TEMP_GS_BUCKET = 'gs://seqr-scratch-temp'


@login_and_policies_required
def save_temp_file(request):

    try:
        uploaded_file_id, filename, json_records = save_uploaded_file(request, allow_json=True)
    except Exception as e:
        return create_json_response({'errors': [str(e)]}, status=400)

    response = {'uploadedFileId': uploaded_file_id}
    if request.GET.get('parsedData'):
        response['parsedData'] = json_records
    else:
        row_summary = 'json' if filename.endswith('.json') else f'{len(json_records)} rows'
        response['info'] = [f'Parsed {row_summary} from {filename}']

    return create_json_response(response)


def _parsed_file_iter(stream, parse_line=lambda l: l):
    for line in stream:
        yield parse_line(line)


def parse_file(filename, stream, iter_file=False, allow_json=False):
    if filename.endswith('.tsv') or filename.endswith('.fam') or filename.endswith('.ped'):
        parse_line = lambda line: [s.strip().strip('"') for s in line.rstrip('\n').split('\t')]
        if iter_file:
            return _parsed_file_iter(stream, parse_line)
        return [parse_line(line) for line in stream]

    elif filename.endswith('.csv'):
        reader = csv.reader(stream)
        if iter_file:
            return _parsed_file_iter(reader)
        return [row for row in reader]

    elif filename.endswith('.xls') or filename.endswith('.xlsx') and not iter_file:
        wb = xl.load_workbook(stream, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[_parse_excel_string_cell(cell) for cell in row] for row in ws.iter_rows()]
        # trim trailing empty rows
        last_row_index = max(i for i, row in enumerate(rows) if any(val for val in row))
        rows = rows[:last_row_index+1]
        # all rows should have same column count
        last_col_index = max(max(i for i, val in enumerate(row) if val) for row in rows)
        padding = [''] * last_col_index
        rows = [(row + padding)[:last_col_index+1] for row in rows]

        return rows

    elif filename.endswith('.json') and allow_json:
        return json.loads(stream.read())

    raise ValueError(f"Unexpected{' iterated' if iter_file else ''} file type: {filename}")


def _parse_excel_string_cell(cell):
    cell_value = cell.value
    if cell_value is not None and cell.data_type == 'n' and int(cell_value) == cell_value:
        cell_value = '{:.0f}'.format(cell_value)
    return cell_value or ''


def get_temp_file_path(file_name, is_local=None):
    if is_local is None:
        is_local = not anvil_enabled()
    if not is_local:
        return f'{TEMP_GS_BUCKET}/{file_name}'

    upload_directory = os.path.join(tempfile.gettempdir(), 'temp_uploads')
    if not os.path.isdir(upload_directory):
        os.makedirs(upload_directory)

    return os.path.join(upload_directory, file_name)


def _compute_serialized_file_name(uploaded_file_id):
    return f'temp_upload_{uploaded_file_id}.json.gz'


def save_uploaded_file(request, process_records=None, allow_json=False):

    if len(request.FILES) != 1:
        raise ValueError("Received %s files instead of 1" % len(request.FILES))

    # parse file
    stream = next(iter(request.FILES.values()))
    filename = stream._name

    if not filename.endswith('.xls') and not filename.endswith('.xlsx'):
        stream = TextIOWrapper(stream.file, encoding = 'utf-8')

    json_records = parse_file(filename, stream, allow_json=allow_json)
    if process_records:
        json_records = process_records(json_records, filename=filename)

    # save json to temporary file
    uploaded_file_id = hashlib.md5(str(json_records).encode('utf-8')).hexdigest() # nosec
    file_name = _compute_serialized_file_name(uploaded_file_id)
    serialized_file_path = get_temp_file_path(file_name, is_local=True)
    with gzip.open(serialized_file_path, 'wt') as f:
        json.dump(json_records, f)

    persist_temp_file(file_name, request.user)

    return uploaded_file_id, filename, json_records


def persist_temp_file(file_name, user):
    if not anvil_enabled():
        return

    src_path = get_temp_file_path(file_name, is_local=True)
    dest_path = get_temp_file_path(file_name, is_local=False)
    mv_file_to_gs(src_path, dest_path, user)


def load_uploaded_file(upload_file_id):
    serialized_file_path = get_temp_file_path(_compute_serialized_file_name(upload_file_id))
    return json.loads(next(file_iter(serialized_file_path)))
