# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from io import StringIO
import mock

import openpyxl as xl
from tempfile import NamedTemporaryFile

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls.base import reverse

from seqr.views.utils.file_utils import save_temp_file, parse_file, load_uploaded_file
from seqr.views.utils.test_utils import AuthenticationTestCase

TSV_DATA = b'Family ID	Individual ID	Notes\n\
"1"	"NA19675"	"An affected individual, additional metadata"\n\
"0"	"NA19678"	""'

CSV_DATA = b'Family ID,Individual ID,Notes\n\
"1","NA19675","An affected individual, additional metadata"\n\
"0","NA19678",""'

JSON_DATA = b'[["Family ID", "Individual ID", "Notes"], ["1", "NA19675", "An affected individual, additional metadata"], ["0", "NA19678", ""]]'

EXCEL_DATA = b'excel data'

TEST_DATA_TYPES = {
    'tsv': TSV_DATA,
    'fam': TSV_DATA,
    'ped': TSV_DATA,
    'csv': CSV_DATA,
    'json': JSON_DATA,
    'xls': EXCEL_DATA,
    'xlsx': EXCEL_DATA,
}

PARSED_DATA = [
    ['Family ID', 'Individual ID', 'Notes'],
    ['1', 'NA19675', 'An affected individual, additional metadata'],
    ['0', 'NA19678', ''],
]


def _mock_cell(value):
    mock_cell = mock.MagicMock()
    mock_cell.value = value
    try:
        mock_cell.value = int(value)
        mock_cell.data_type = 'n'
    except ValueError:
        mock_cell.data_type = 's'
    return mock_cell


MOCK_EXCEL_SHEET = mock.MagicMock()
MOCK_EXCEL_SHEET.iter_rows.return_value = [[_mock_cell(cell) for cell in row] for row in PARSED_DATA]


class FileUtilsTest(AuthenticationTestCase):
    fixtures = ['users']

    def setUp(self):
        # Test uploading with xlsx data
        wb = xl.Workbook()
        ws = wb[wb.sheetnames[0]]
        ws['A1'], ws['B1'], ws['C1'] = ['Family ID', 'Individual ID', 'Notes']
        ws['A2'], ws['B2'], ws['C2'] = [1, 'NA19675', 'An affected individual, additional metadata']
        ws['A3'], ws['B3']           = [0, 'NA19678']
        ws['A4']                     = '' # for testing trimming trailing empty rows

        with NamedTemporaryFile() as tmp:
            wb.save(tmp)
            tmp.seek(0)
            self.xlsx_data = tmp.read()

    def test_temp_file_upload(self):
        url = reverse(save_temp_file)
        self.check_require_login(url)

        response = self.client.post(url, {
            'f': SimpleUploadedFile("test_data.tsv", TSV_DATA),
            'invalid': SimpleUploadedFile("test_data.foo", TSV_DATA),
        })
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'errors': ['Received 2 files instead of 1']})

        response = self.client.post(url, {'invalid': SimpleUploadedFile("test_data.foo", TSV_DATA)})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'errors': ['Unexpected file type: test_data.foo']})

        response = self.client.post(url, {'f': SimpleUploadedFile("test_data.tsv", TSV_DATA)})
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertDictEqual(response_json, {
            'info': ['Parsed 3 rows from test_data.tsv'],
            'uploadedFileId': mock.ANY,
        })

        # Test loading uploaded file
        uploaded_file_id = response_json['uploadedFileId']
        file_content = load_uploaded_file(uploaded_file_id)
        self.assertListEqual(file_content, PARSED_DATA)
        # File should be removed after loading it once
        with self.assertRaises(IOError):
            load_uploaded_file(uploaded_file_id)

        # Test uploading with returned data and test with file formats other than 'xls' and 'xlsx'
        for ext, data in TEST_DATA_TYPES.items():
            if ext == 'xls' or ext == 'xlsx':
                data = self.xlsx_data
            response = self.client.post(
                '{}?parsedData=true'.format(url), {'f': SimpleUploadedFile("test_data.{}".format(ext), data)})
            self.assertEqual(response.status_code, 200)
            self.assertDictEqual(response.json(), {
                'parsedData': PARSED_DATA,
                'uploadedFileId': mock.ANY,
            })

    @mock.patch('seqr.views.utils.file_utils.xl.load_workbook')
    def test_parse_file(self, mock_load_xl):
        mock_load_xl.return_value.sheetnames = ['sheet1']
        mock_load_xl.return_value.__getitem__.return_value = MOCK_EXCEL_SHEET

        for ext, data in TEST_DATA_TYPES.items():
            self.assertListEqual(parse_file('test.{}'.format(ext), StringIO(data.decode('utf-8'))), PARSED_DATA)

        for call_args in mock_load_xl.call_args_list:
            self.assertEqual(call_args.args[0].read().encode('utf-8'), EXCEL_DATA)
            self.assertDictEqual(call_args.kwargs, {'read_only': True})
