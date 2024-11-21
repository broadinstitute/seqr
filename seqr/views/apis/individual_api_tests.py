# -*- coding: utf-8 -*-
import datetime
import gzip
import json
import mock
import re

from copy import deepcopy
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls.base import reverse
from io import BytesIO
from openpyxl import load_workbook

from seqr.models import Individual, Sample, SavedVariant, VariantTag
from seqr.views.apis.individual_api import edit_individuals_handler, update_individual_handler, \
    delete_individuals_handler, receive_individuals_table_handler, save_individuals_table_handler, \
    receive_individuals_metadata_handler, save_individuals_metadata_table_handler, update_individual_hpo_terms, \
    get_hpo_terms, get_individual_rna_seq_data, import_gregor_metadata
from seqr.views.apis.report_api_tests import PARTICIPANT_TABLE, PHENOTYPE_TABLE, EXPERIMENT_TABLE, EXPERIMENT_LOOKUP_TABLE, GENETIC_FINDINGS_TABLE
from seqr.views.utils.test_utils import AuthenticationTestCase, AnvilAuthenticationTestCase, INDIVIDUAL_FIELDS, \
    INDIVIDUAL_CORE_FIELDS, CORE_INTERNAL_INDIVIDUAL_FIELDS, GENE_FIELDS

PROJECT_GUID = 'R0001_1kg'
PM_REQUIRED_PROJECT_GUID = 'R0003_test'
EXTERNAL_WORKSPACE_PROJECT_GUID = 'R0004_non_analyst_project'

INDIVIDUAL_GUID = "I000001_na19675"
ID_UPDATE_GUID = "I000002_na19678"
UPDATED_ID = "NA19678_1"
UPDATED_MATERNAL_ID = "NA20870"

INDIVIDUAL_IDS_UPDATE_DATA = {
    'individualGuid': ID_UPDATE_GUID,
    'familyId': '1',
    'individualId': UPDATED_ID,
    'maternalId': UPDATED_MATERNAL_ID,
    'paternalId': '',
    'sex': 'M',
    'notes': '',
    'population': None,
    'filterFlags': None,
    'projectGuid': PROJECT_GUID,
}

INDIVIDUAL_UPDATE_GUID = "I000007_na20870"
INDIVIDUAL_UPDATE_DATA = {
    'displayName': 'NA20870',
    'notes': 'A note',
    'birthYear': 2000,
    'features': [{
        'id': 'HP:0002011',
        'label': 'nervous system abnormality',
    }, {
        'id': 'HP:0002011',
        'label': 'nervous system abnormality',
        'category': 'HP:0000708',
        'categoryName': 'Nervous',
        'qualifiers': [{'type': 'onset', 'label': 'congenital'}],
    }, {
        'id': 'HP:0011675',
        'notes': 'A new term',
    }],
    'absentFeatures': [],
    'absentNonstandardFeatures': [{'id': 'Some new feature', 'notes': 'No term for this', 'details': 'extra detail'}]
}

PM_REQUIRED_INDIVIDUAL_GUID = 'I000017_na20889'
PM_REQUIRED_INDIVIDUAL_UPDATE_DATA = {
    'individualGuid': PM_REQUIRED_INDIVIDUAL_GUID, 'individualId': 'NA20889', 'familyId': '12', 'displayName': 'NA20889_a'
}

EXTERNAL_WORKSPACE_INDIVIDUAL_GUID = 'I000019_na21987'
EXTERNAL_WORKSPACE_INDIVIDUAL_UPDATE_DATA = {
    'individualGuid': EXTERNAL_WORKSPACE_INDIVIDUAL_GUID,
    'individualId': 'NA21987',
    'paternalGuid': 'I000018_na21234',
    'maternalGuid': 'I000020_na65432',
    'maternalId': '',
    'paternalId': 'foobar',
    'sex': 'U',
    'affected': 'N',
}

FAMILY_UPDATE_GUID = "I000007_na20870"
INDIVIDUAL_FAMILY_UPDATE_DATA = {
    "individualGuid": FAMILY_UPDATE_GUID,
    "familyId": "1",
    "individualId": UPDATED_MATERNAL_ID,
}

LOAD_PARTICIPANT_TABLE = deepcopy(PARTICIPANT_TABLE)
for row in LOAD_PARTICIPANT_TABLE[4:]:
    row[7] = row[7].replace('Broad_', '')
LOAD_PARTICIPANT_TABLE[6][15] += '|Asian'
LOAD_PARTICIPANT_TABLE[6][17] = ''


@mock.patch('seqr.utils.middleware.DEBUG', False)
class IndividualAPITest(object):

    def test_update_individual_handler(self):
        edit_individuals_url = reverse(update_individual_handler, args=[INDIVIDUAL_UPDATE_GUID])
        self.check_collaborator_login(edit_individuals_url)

        response = self.client.post(edit_individuals_url, content_type='application/json',
                                    data=json.dumps(INDIVIDUAL_UPDATE_DATA))

        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertListEqual(list(response_json.keys()), [INDIVIDUAL_UPDATE_GUID])
        self.assertSetEqual(set(response_json[INDIVIDUAL_UPDATE_GUID].keys()), INDIVIDUAL_CORE_FIELDS)
        individual = Individual.objects.get(guid=INDIVIDUAL_UPDATE_GUID)
        self.assertEqual(response_json[INDIVIDUAL_UPDATE_GUID]['displayName'], 'NA20870')
        self.assertEqual(individual.display_name, '')
        self.assertEqual(response_json[INDIVIDUAL_UPDATE_GUID]['notes'], 'A note')
        self.assertIsNone(response_json[INDIVIDUAL_UPDATE_GUID]['birthYear'])
        self.assertFalse('features' in response_json[INDIVIDUAL_UPDATE_GUID])
        self.assertIsNone(individual.features)

        self.login_manager()
        response = self.client.post(edit_individuals_url, content_type='application/json',
                                    data=json.dumps(INDIVIDUAL_UPDATE_DATA))
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json[INDIVIDUAL_UPDATE_GUID].keys()), INDIVIDUAL_CORE_FIELDS)
        self.assertEqual(response_json[INDIVIDUAL_UPDATE_GUID]['birthYear'], 2000)

        update_json = {'analyteType': 'D', 'tissueAffectedStatus': False}
        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps(update_json))
        self.assertEqual(response.status_code, 403)

        self.login_analyst_user()
        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps(update_json))
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        response_fields = deepcopy(INDIVIDUAL_CORE_FIELDS)
        response_fields.update(CORE_INTERNAL_INDIVIDUAL_FIELDS)
        self.assertSetEqual(set(response_json[INDIVIDUAL_UPDATE_GUID].keys()), response_fields)
        self.assertEqual(response_json[INDIVIDUAL_UPDATE_GUID]['analyteType'], 'D')
        self.assertFalse(response_json[INDIVIDUAL_UPDATE_GUID]['tissueAffectedStatus'])

    def test_update_individual_hpo_terms(self):
        edit_individuals_url = reverse(update_individual_hpo_terms, args=[INDIVIDUAL_UPDATE_GUID])
        self.check_manager_login(edit_individuals_url)

        response = self.client.post(edit_individuals_url, content_type='application/json',
                                    data=json.dumps(INDIVIDUAL_UPDATE_DATA))

        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertListEqual(list(response_json.keys()), [INDIVIDUAL_UPDATE_GUID])
        self.assertListEqual(response_json[INDIVIDUAL_UPDATE_GUID]['features'], [
            {
                'id': 'HP:0002011',
                'category': 'HP:0000707',
                'label': 'Morphological abnormality of the central nervous system',
                'qualifiers': [{'type': 'onset', 'label': 'congenital'}],
            },
            {'id': 'HP:0011675', 'category': 'HP:0001626', 'label': 'Arrhythmia', 'notes': 'A new term'},
        ])
        self.assertListEqual(response_json[INDIVIDUAL_UPDATE_GUID]['absentNonstandardFeatures'], [
            {'id': 'Some new feature', 'notes': 'No term for this'}
        ])
        self.assertIsNone(response_json[INDIVIDUAL_UPDATE_GUID]['absentFeatures'])
        self.assertIsNone(response_json[INDIVIDUAL_UPDATE_GUID]['nonstandardFeatures'])

        self.assertListEqual(Individual.objects.get(guid=INDIVIDUAL_UPDATE_GUID).features, [
            {'id': 'HP:0002011', 'qualifiers': [{'type': 'onset', 'label': 'congenital'}]},
            {'id': 'HP:0011675', 'notes': 'A new term'},
        ])

    @mock.patch('seqr.views.utils.permissions_utils.PM_USER_GROUP')
    def test_edit_individuals(self, mock_pm_group):
        edit_individuals_url = reverse(edit_individuals_handler, args=[PROJECT_GUID])
        self.check_manager_login(edit_individuals_url)

        # send invalid requests
        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps({}))
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.reason_phrase, "'individuals' not specified")

        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [INDIVIDUAL_IDS_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 400)
        self.assertListEqual(response.json()['errors'], [
            'NA19678 already has loaded data and cannot update the ID',
            "NA20870 is the mother of NA19678_1 but is not included. Make sure to create an additional record with NA20870 as the Individual ID",
        ])

        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [INDIVIDUAL_IDS_UPDATE_DATA, INDIVIDUAL_FAMILY_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 400)
        self.assertListEqual(response.json()['errors'], [
            'NA19678 already has loaded data and cannot update the ID',
            'NA20870 already has loaded data and cannot be moved to a different family',
        ])

        # send valid request
        Sample.objects.filter(guid__in=['S000130_na19678', 'S000135_na20870']).update(is_active=False)
        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [INDIVIDUAL_IDS_UPDATE_DATA, INDIVIDUAL_FAMILY_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 200)
        response_json = response.json()

        self.assertSetEqual(set(response_json.keys()), {'individualsByGuid', 'familiesByGuid'})
        self.assertSetEqual({'F000001_1', 'F000003_3'}, set(response_json['familiesByGuid']))
        self.assertSetEqual({ID_UPDATE_GUID, FAMILY_UPDATE_GUID, INDIVIDUAL_GUID, "I000003_na19679"},
                            set(response_json['familiesByGuid']['F000001_1']['individualGuids']))
        self.assertListEqual(response_json['familiesByGuid']['F000003_3']['individualGuids'], [])
        self.assertIsNone(response_json['familiesByGuid']['F000001_1']['pedigreeImage'])
        self.assertIsNone(response_json['familiesByGuid']['F000003_3']['pedigreeImage'])

        self.assertSetEqual({ID_UPDATE_GUID, FAMILY_UPDATE_GUID, INDIVIDUAL_GUID},
                            set(response_json['individualsByGuid']))
        self.assertEqual(response_json['individualsByGuid'][ID_UPDATE_GUID]['individualId'], UPDATED_ID)
        self.assertEqual(response_json['individualsByGuid'][ID_UPDATE_GUID]['maternalId'], UPDATED_MATERNAL_ID)
        self.assertEqual(response_json['individualsByGuid'][INDIVIDUAL_GUID]['paternalId'], UPDATED_ID)

        # test only updating parental IDs
        response = self.client.post(edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [{
                'individualGuid': ID_UPDATE_GUID,
                'familyId': '1',
                'individualId': UPDATED_ID,
                'maternalId': 'NA19679',
            }]
        }))
        self.assertEqual(response.status_code, 200)
        response_json = response.json()

        self.assertSetEqual(set(response_json.keys()), {'individualsByGuid', 'familiesByGuid'})
        self.assertDictEqual({}, response_json['familiesByGuid'])
        self.assertSetEqual({ID_UPDATE_GUID}, set(response_json['individualsByGuid']))
        self.assertEqual(response_json['individualsByGuid'][ID_UPDATE_GUID]['individualId'], UPDATED_ID)
        self.assertEqual(response_json['individualsByGuid'][ID_UPDATE_GUID]['maternalId'], 'NA19679')

        # Test PM permission
        pm_required_edit_individuals_url = reverse(edit_individuals_handler, args=[PM_REQUIRED_PROJECT_GUID])
        response = self.client.post(pm_required_edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [PM_REQUIRED_INDIVIDUAL_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 403)

        self.login_pm_user()
        response = self.client.post(pm_required_edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [PM_REQUIRED_INDIVIDUAL_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 403)

        mock_pm_group.__bool__.return_value = True
        mock_pm_group.resolve_expression.return_value = 'project-managers'
        mock_pm_group.__eq__.side_effect = lambda s: s == 'project-managers'
        response = self.client.post(pm_required_edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [PM_REQUIRED_INDIVIDUAL_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 200)
        self.assertDictEqual(response.json(), {
            'individualsByGuid': {PM_REQUIRED_INDIVIDUAL_GUID: mock.ANY},
            'familiesByGuid': {}
        })

        # Test External AnVIL projects
        ext_anvil_edit_individuals_url = reverse(edit_individuals_handler, args=[EXTERNAL_WORKSPACE_PROJECT_GUID])
        self.login_collaborator()
        response = self.client.post(
            ext_anvil_edit_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [EXTERNAL_WORKSPACE_INDIVIDUAL_UPDATE_DATA]
            }))
        self.assertEqual(response.status_code, 403)

        self.login_manager()
        response = self.client.post(
            ext_anvil_edit_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [EXTERNAL_WORKSPACE_INDIVIDUAL_UPDATE_DATA]
            }))

        if not self.HAS_EXTERNAL_PROJECT_ACCESS:
            self.assertEqual(response.status_code, 403)
            return

        self.assertEqual(response.status_code, 400)
        self.assertListEqual(response.json()['errors'], [
            'Invalid parental guid I000020_na65432',
            'NA21234 is recorded as Female sex and also as the father of NA21987',
        ])

        update_json = deepcopy(EXTERNAL_WORKSPACE_INDIVIDUAL_UPDATE_DATA)
        update_json['maternalGuid'] = update_json.pop('paternalGuid')
        response = self.client.post(ext_anvil_edit_individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [update_json]
        }))

        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertDictEqual(response_json, {
            'individualsByGuid': {EXTERNAL_WORKSPACE_INDIVIDUAL_GUID: mock.ANY},
            'familiesByGuid': {}
        })
        updated_individual = response_json['individualsByGuid'][EXTERNAL_WORKSPACE_INDIVIDUAL_GUID]
        self.assertEqual(updated_individual['sex'], 'U')
        self.assertEqual(updated_individual['affected'], 'N')
        self.assertEqual(updated_individual['maternalGuid'], 'I000018_na21234')
        self.assertEqual(updated_individual['maternalId'], 'NA21234')
        self.assertIsNone(updated_individual['paternalGuid'])
        self.assertIsNone(updated_individual['paternalGuid'])

    @mock.patch('seqr.views.utils.permissions_utils.PM_USER_GROUP')
    def test_delete_individuals(self, mock_pm_group):
        individuals_url = reverse(delete_individuals_handler, args=[PROJECT_GUID])
        self.check_manager_login(individuals_url)

        # send invalid requests
        response = self.client.post(individuals_url, content_type='application/json', data=json.dumps({
            'individualsX': [INDIVIDUAL_IDS_UPDATE_DATA]
        }))
        self.assertEqual(response.status_code, 400)

        # send valid requests
        response = self.client.post(individuals_url, content_type='application/json', data=json.dumps({
            'individuals': [INDIVIDUAL_IDS_UPDATE_DATA]
        }))
        self._assert_expected_delete_individuals(response, mock_pm_group)

    def _assert_expected_delete_individuals(self, response, mock_pm_group):
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {'individualsByGuid', 'familiesByGuid'})
        self.assertDictEqual(response_json, {
            'individualsByGuid': {'I000002_na19678': None},
            'familiesByGuid': {'F000001_1': mock.ANY}
        })
        self.assertFalse('I000002_na19678' in response_json['familiesByGuid']['F000001_1']['individualGuids'])
        self.assertIsNone(response_json['familiesByGuid']['F000001_1']['pedigreeImage'])

        # Test PM permission
        pm_required_delete_individuals_url = reverse(delete_individuals_handler, args=[PM_REQUIRED_PROJECT_GUID])
        response = self.client.post(
            pm_required_delete_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [PM_REQUIRED_INDIVIDUAL_UPDATE_DATA]
            }))
        self.assertEqual(response.status_code, 403)

        self.login_pm_user()
        response = self.client.post(
            pm_required_delete_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [PM_REQUIRED_INDIVIDUAL_UPDATE_DATA]
            }))
        self.assertEqual(response.status_code, 403)

        mock_pm_group.__bool__.return_value = True
        mock_pm_group.resolve_expression.return_value = 'project-managers'
        mock_pm_group.__eq__.side_effect = lambda s: s == 'project-managers'
        response = self.client.post(
            pm_required_delete_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [PM_REQUIRED_INDIVIDUAL_UPDATE_DATA]
            }))

        self.assertEqual(response.status_code, 400)
        self.assertListEqual(response.json()['errors'], ['Unable to delete individuals with active MME submission: NA20889'])

        data = json.dumps({
            'individuals': [{'individualGuid': 'I000015_na20885'}]
        })
        response = self.client.post(pm_required_delete_individuals_url, content_type='application/json', data=data)
        self.assertEqual(response.status_code, 200)

        # Test External AnVIL projects
        ext_anvil_delete_individuals_url = reverse(delete_individuals_handler, args=[EXTERNAL_WORKSPACE_PROJECT_GUID])
        self.login_collaborator()
        response = self.client.post(
            ext_anvil_delete_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [EXTERNAL_WORKSPACE_INDIVIDUAL_UPDATE_DATA]
            }))
        self.assertEqual(response.status_code, 403)

        self.login_manager()
        response = self.client.post(
            ext_anvil_delete_individuals_url, content_type='application/json', data=json.dumps({
                'individuals': [EXTERNAL_WORKSPACE_INDIVIDUAL_UPDATE_DATA]
            }))
        self.assertEqual(response.status_code, 200 if self.HAS_EXTERNAL_PROJECT_ACCESS else 403)

    def test_individuals_table_handler_errors(self):
        individuals_url = reverse(receive_individuals_table_handler, args=[PROJECT_GUID])
        self.check_manager_login(individuals_url)

        response = self.client.get(individuals_url)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'errors': ['Received 0 files instead of 1'], 'warnings': []})

        response = self.client.post(individuals_url, {'f': SimpleUploadedFile('test.tsv', 'family   indiv\n1    '.encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'errors': mock.ANY, 'warnings': None})
        errors = response.json()['errors']
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0], 'Missing required columns: Individual Id')

        response = self.client.post(individuals_url, {'f': SimpleUploadedFile(
            'test.tsv', 'Family ID	Individual ID\n""	""""'.encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': None, 'errors': [
            'Missing Family Id in row #1', 'Missing Individual Id in row #1',
        ]})

        response = self.client.post(individuals_url, {'f': SimpleUploadedFile(
            'test.tsv',  '#Some comments\n#Family ID	#Individual ID	Previous Individual ID\n"1"	"NA19675_1"""'.encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': [], 'errors': [
            'Error while parsing file: test.tsv. Row 1 contains 2 columns: 1, NA19675_1, while header contains 3: Family ID, Individual ID, Previous Individual ID',
        ]})

        response = self.client.post(individuals_url, {'f': SimpleUploadedFile(
            'test.tsv', 'Family ID	Individual ID	Previous Individual ID\n"1"	"NA19675_1"	"NA19675"\n"2"	"NA19675_1"	""'.encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {
            'errors': [
                'NA19675_1 already has loaded data and cannot be moved to a different family',
                'NA19675_1 is included as 2 separate records, but must be unique within the project',
            ], 'warnings': []
        })

        response = self.client.post(individuals_url, {'f': SimpleUploadedFile(
            'test.tsv', 'Family ID	Individual ID	Previous Individual ID\n"1"	"NA19675_1"	"NA19675"'.encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {
            'errors': ['Could not find individuals with the following previous IDs: NA19675'], 'warnings': []
        })

        response = self.client.post(individuals_url, {'f': SimpleUploadedFile(
            'test.tsv', 'Family ID	Individual ID	affected	sex	proband_relation\n"1"	"NA19675_1"	"no"	"boy"	"mom"'.encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': None, 'errors': [
            'Invalid value "no" for Affected in row #1',
            'Invalid value "boy" for Sex in row #1',
            'Invalid value "mom" for Proband Relationship in row #1',
        ]})

        rows = [
            'Family ID	Individual ID	Paternal ID	sex	proband_relation	affected',
            '"1"	"NA19675_1"	"NA19678_dad"	""	""	"affect"',
        ]
        response = self.client.post(individuals_url, {
            'f': SimpleUploadedFile('test.tsv',  '\n'.join(rows).encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        missing_entry_warning = "NA19678_dad is the father of NA19675_1 but is not included. Make sure to create an additional record with NA19678_dad as the Individual ID"
        self.assertDictEqual(response.json(), {
            'errors': [missing_entry_warning],
            'warnings': [],
        })

        rows += [
            '"1"	"NA19675_1"	"NA19675_1"	"F"	"Father"	"unaffected"',
            '"2"	"NA19675_2"	"NA19675_1"	"XXX"	"Nephew"	"unknown"',
            '"2"	"NA19677"	"NA19675_2"	"M"	""	"unaffected"',
        ]
        response = self.client.post(individuals_url, {
            'f': SimpleUploadedFile('test.tsv', '\n'.join(rows).encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {
            'errors': [
                'Invalid proband relationship "Father" for NA19675_1 with given gender Female',
                'NA19675_1 is recorded as their own father',
                'NA19675_1 is recorded as Female sex and also as the father of NA19675_1',
                'Invalid proband relationship "Nephew" for NA19675_2 with given gender XXX',
                'NA19675_1 is recorded as Female sex and also as the father of NA19675_2',
                'NA19675_1 is recorded as the father of NA19675_2 but they have different family ids: 1 and 2',
                'NA19675_2 is recorded as XXX sex and also as the father of NA19677',
                'NA19675_1 is included as 2 separate records, but must be unique within the project',
            ],
            'warnings': [missing_entry_warning, 'The following families do not have any affected individuals: 2'],
        })

        rows = [rows[0], '"new_fam_1"	"NA19677"	""	"M"	""	"unaffected"']
        response = self.client.post(individuals_url, {
            'f': SimpleUploadedFile('test.tsv', '\n'.join(rows).encode('utf-8'))})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'errors': [
            'The following families do not have any affected individuals: new_fam_1'
        ], 'warnings': []})

    @mock.patch('seqr.views.utils.permissions_utils.PM_USER_GROUP', 'project-managers')
    def test_individuals_table_handler(self):
        individuals_url = reverse(receive_individuals_table_handler, args=[PROJECT_GUID])
        self.check_manager_login(individuals_url)

        data = 'Family ID	Individual ID	Previous Individual ID	Paternal ID	Maternal ID	Sex	Affected Status	Notes	familyNotes\n\
"1"	" NA19675_1 "	""	"NA19678 "	"NA19679"	"Female"	"Affected"	"A affected individual, test1-zsf"	""\n\
"1"	"NA19678"	""	""	""	"XXY"	"Unaffected"	"a individual note"	""\n\
"4"	"NA20872_update"	"NA20872"	""	""	"Male"	"Affected"	""	""\n\
"21"	" HG00735"	""	""	""	"Female"	"Affected"	""	"a new family""'

        f = SimpleUploadedFile("1000_genomes demo_individuals.tsv", data.encode('utf-8'))

        response = self.client.post(individuals_url, {'f': f})
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {'info', 'errors', 'warnings', 'uploadedFileId'})
        self.assertListEqual(response_json['errors'], [])
        self.assertListEqual(response_json['warnings'], [])
        self.assertListEqual(response_json['info'], [
            '3 families, 4 individuals parsed from 1000_genomes demo_individuals.tsv',
            '1 new families, 1 new individuals will be added to the project',
            '3 existing individuals will be updated',
        ])

        url = reverse(save_individuals_table_handler, args=[PROJECT_GUID, response_json['uploadedFileId']])

        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {'individualsByGuid', 'familiesByGuid', 'familyNotesByGuid'})

        self.assertEqual(len(response_json['familiesByGuid']), 3)
        self.assertTrue('F000001_1' in response_json['familiesByGuid'])
        self.assertTrue('F000004_4' in response_json['familiesByGuid'])
        new_family_guid = next(guid for guid in response_json['familiesByGuid'].keys() if guid != 'F000001_1' and guid != 'F000004_4')
        self.assertEqual(response_json['familiesByGuid'][new_family_guid]['familyId'], '21')
        self.assertIsNone(response_json['familiesByGuid']['F000001_1']['pedigreeImage'])

        self.assertEqual(len(response_json['familyNotesByGuid']), 1)
        new_note = list(response_json['familyNotesByGuid'].values())[0]
        self.assertEqual(new_note['note'], 'a new family')
        self.assertEqual(new_note['noteType'], 'C')
        self.assertEqual(new_note['createdBy'], 'Test Manager User')

        self.assertEqual(len(response_json['individualsByGuid']), 4)
        self.assertTrue('I000001_na19675' in response_json['individualsByGuid'])
        self.assertTrue('I000002_na19678' in response_json['individualsByGuid'])
        self.assertTrue('I000008_na20872' in response_json['individualsByGuid'])
        new_indiv_guid = next(guid for guid in response_json['individualsByGuid'].keys()
                              if guid not in {'I000001_na19675', 'I000002_na19678', 'I000008_na20872'})
        self.assertEqual(response_json['individualsByGuid']['I000001_na19675']['individualId'], 'NA19675_1')
        self.assertEqual(response_json['individualsByGuid']['I000001_na19675']['sex'], 'F')
        self.assertEqual(
            response_json['individualsByGuid']['I000001_na19675']['notes'], 'A affected individual, test1-zsf')
        self.assertEqual(response_json['individualsByGuid']['I000002_na19678']['sex'], 'XXY')
        self.assertEqual(response_json['individualsByGuid'][new_indiv_guid]['individualId'], 'HG00735')
        self.assertEqual(response_json['individualsByGuid'][new_indiv_guid]['sex'], 'F')
        self.assertEqual(response_json['individualsByGuid']['I000008_na20872']['individualId'], 'NA20872_update')

        # Test PM permission
        receive_url = reverse(receive_individuals_table_handler, args=[PM_REQUIRED_PROJECT_GUID])
        save_url = reverse(save_individuals_table_handler, args=[PM_REQUIRED_PROJECT_GUID, '123'])
        response = self.client.post(receive_url, {'f': f})
        self.assertEqual(response.status_code, 403)
        response = self.client.post(save_url)
        self.assertEqual(response.status_code, 403)

        self.login_pm_user()
        response = self.client.post(receive_url, {
            'f': SimpleUploadedFile('individuals.tsv', 'Family ID	Individual ID\n1	2'.encode('utf-8'))})
        self.assertEqual(response.status_code, 200)
        save_url = reverse(save_individuals_table_handler, args=[
            PM_REQUIRED_PROJECT_GUID, response.json()['uploadedFileId']])
        response = self.client.post(save_url)
        self.assertEqual(response.status_code, 200)

    @mock.patch('seqr.views.utils.permissions_utils.PM_USER_GROUP', 'project-managers')
    @mock.patch('seqr.views.utils.pedigree_info_utils.NO_VALIDATE_MANIFEST_PROJECT_CATEGORIES')
    @mock.patch('seqr.utils.communication_utils.EmailMultiAlternatives')
    def test_individuals_sample_manifest_table_handler(self, mock_email, mock_no_validate_categories):
        receive_url = reverse(receive_individuals_table_handler, args=[PROJECT_GUID])
        self.check_manager_login(receive_url)

        def _send_request_data(data):
            return self.client.post(receive_url, {'f': SimpleUploadedFile(
                'sample_manifest.tsv', '\n'.join(['\t'.join([str(c) for c in row]) for row in data]).encode('utf-8')),
            })

        header_2 = [
            'Kit ID', 'Well', 'Sample ID', 'Family ID', 'Alias', 'Alias', 'Paternal Sample ID', 'Maternal Sample ID',
            'Gender', 'Affected Status', 'Primary Biosample', 'Analyte Type', 'Tissue Affected Status', 'Recontactable',
            'Volume', 'Concentration', 'Notes', 'MONDO Label', 'MONDO ID', 'Consent Code', 'Data Use Restrictions']
        header_3 = [
            '', 'Position', '', '', 'Collaborator Participant ID', 'Collaborator Sample ID', '', '', '', '', '', '',
            '(i.e yes, no)', '(i.e yes, no, unknown)', 'ul', 'ng/ul', '', '', '(i.e. "MONDO:0031632")', '',
            'indicate study/protocol number']
        data = [
            ['Do not modify - Broad use', '', '', 'Please fill in columns D - T', '', '', '', '', '', '', '', '', '',
             '', '', '', '', '', '', '', ''],
            header_2, header_3,
        ]

        response = _send_request_data(data)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {
            'errors': ['Unsupported file format'], 'warnings': [],
        })

        data[1] = header_2[:5] + header_2[7:10] + header_2[14:17] + ['Coded Phenotype'] + header_2[19:]
        response = _send_request_data(data)
        self.assertDictEqual(response.json(), {
            'errors': ['Unsupported file format'], 'warnings': [],
        })

        self.login_pm_user()
        response = _send_request_data(data)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': [], 'errors': [
            'Expected vs. actual header columns: | '
            'Sample ID| Family ID| Alias|-Alias|-Paternal Sample ID| Maternal Sample ID| Gender| Affected Status|'
            '-Primary Biosample|-Analyte Type|-Tissue Affected Status|-Recontactable| Volume| Concentration| Notes|'
            '-MONDO Label|-MONDO ID|+Coded Phenotype| Consent Code| Data Use Restrictions',
        ]})

        data[1] = header_2
        data[2] = header_3[:4] + header_3[5:10] + header_3[14:18] + header_3[-1:]
        response = _send_request_data(data)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': [], 'errors': [
            'Expected vs. actual header columns: |-Collaborator Participant ID| Collaborator Sample ID|+',
        ]})

        data[2] = header_3
        data += [
            ['SK-3QVD', 'A02', 'SM-IRW6C', 'PED073', 'SCO_PED073B_GA0339', 'SCO_PED073B_GA0339_1', '', '', 'male',
             'unaffected', 'UBERON:0000479 (tissue)', 'blood plasma', '', 'Unknown', '20', 94.8, 'probably dad', '',
             '', 'GMB', '1234'],
            ['SK-3QVD', 'A03', 'SM-IRW69', 'PED073', 'SCO_PED073C_GA0340', 'SCO_PED073C_GA0340_1',
             'SCO_PED073B_GA0339_1', 'SCO_PED073A_GA0338_1', 'female', 'affected', 'UBERON:0002371 (bone marrow)',
             'DNA', 'Yes', 'No', '20', '98', '', 'Perinatal death', 'MONDO:0100086', 'HMB', '',
             ],
            ['SK-3QVD', 'A04', 'SM-IRW61', 'PED073', 'SCO_PED073C_GA0341', 'SCO_PED073C_GA0341_1',
             'SCO_PED073B_GA0339_1', '', 'male', 'affected', 'UBERON:0002371 (bone marrow)',
             'RNA', 'No', 'No', '17', '83', 'half sib', 'Perinatal death', 'MONDO:0100086', '', '',
             ]]

        expected_warning = 'SCO_PED073A_GA0338_1 is the mother of SCO_PED073C_GA0340_1 but is not included. ' \
                           'Make sure to create an additional record with SCO_PED073A_GA0338_1 as the Individual ID'
        missing_columns_error = 'SCO_PED073B_GA0339_1 is missing the following required columns: MONDO ID, MONDO Label, Tissue Affected Status'
        response = _send_request_data(data)
        self.assertDictEqual(response.json(), {'warnings': [expected_warning], 'errors': [
            missing_columns_error, 'Multiple consent codes specified in manifest: GMB, HMB',
        ]})

        data[4][-2] = 'GMB'
        mock_no_validate_categories.resolve_expression.return_value = ['Not-used category']
        response = _send_request_data(data)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': [expected_warning], 'errors': [
            missing_columns_error, 'Consent code in manifest "GMB" does not match project consent code "HMB"',
        ]})

        data[3][12] = 'Maybe'
        response = _send_request_data(data)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': None, 'errors': ['Invalid value "Maybe" for Tissue Affected Status in row #1']})

        data[3][12] = 'Unknown'
        data[3][17] = 'microcephaly'
        data[3][18] = 'MONDO:0001149'
        data[3][-2] = ''
        data[4][-2] = 'HMB'
        response = _send_request_data(data)
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'warnings': [], 'errors': [expected_warning]})

        data[4][7] = ''
        response = _send_request_data(data)
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertDictEqual(response_json, {'uploadedFileId': mock.ANY, 'warnings': [], 'errors': [], 'info': [
            '1 families, 3 individuals parsed from sample_manifest.tsv',
            '1 new families, 3 new individuals will be added to the project',
            '0 existing individuals will be updated',
        ]})

        mock_email.assert_called_with(
            subject='SK-3QVD Merged Sample Pedigree File',
            body=mock.ANY,
            to=['test_pm_user@test.com'],
            attachments=[
                ('SK-3QVD.xlsx', mock.ANY,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ('sample_manifest.xlsx', mock.ANY,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ])
        self.assertEqual(
            mock_email.call_args.kwargs['body'],
            '\n'.join([
                'User test_pm_user@test.com just uploaded pedigree info to 1kg project n\xe5me with uni\xe7\xf8de.',
                'This email has 2 attached files:',
                '    ', '    SK-3QVD.xlsx is the sample manifest file in a format that can be sent to GP.', '    ',
                '    sample_manifest.tsv is the original merged pedigree-sample-manifest file that the user uploaded.', '    ',
            ]))
        mock_email.return_value.attach_alternative.assert_called_with(
            """User test_pm_user@test.com just uploaded pedigree info to 1kg project n\xe5me with uni\xe7\xf8de.<br />This email has 2 attached files:<br />
    <br />
    <b>SK-3QVD.xlsx</b> is the sample manifest file in a format that can be sent to GP.<br />
    <br />
    <b>sample_manifest.tsv</b> is the original merged pedigree-sample-manifest file that the user uploaded.<br />
    """.replace('\n', ''), 'text/html')
        mock_email.return_value.send.assert_called()

        # Test sent sample manifest is correct
        sample_wb = load_workbook(BytesIO(mock_email.call_args.kwargs['attachments'][0][1]))
        sample_ws = sample_wb.active
        sample_ws.title = 'Sample Info'
        self.assertListEqual(
            [[cell.value or '' for cell in row] for row in sample_ws],
            [['Well', 'Sample ID', 'Alias', 'Alias', 'Gender', 'Volume', 'Concentration'],
             ['Position', '', 'Collaborator Participant ID', 'Collaborator Sample ID', '', 'ul', 'ng/ul'],
             ['A02', 'SM-IRW6C', 'SCO_PED073B_GA0339', 'SCO_PED073B_GA0339_1', 'male', '20', '94.8'],
             ['A03', 'SM-IRW69', 'SCO_PED073C_GA0340', 'SCO_PED073C_GA0340_1', 'female', '20', '98'],
             ['A04', 'SM-IRW61', 'SCO_PED073C_GA0341', 'SCO_PED073C_GA0341_1', 'male', '17', '83']])

        # Test original file copy is correct
        original_wb = load_workbook(BytesIO(mock_email.call_args.kwargs['attachments'][1][1]))
        original_ws = original_wb.active
        self.assertListEqual([[cell.value or '' for cell in row] for row in original_ws], [[str(c) for c in row] for row in data])

        url = reverse(save_individuals_table_handler, args=[PROJECT_GUID, response_json['uploadedFileId']])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {'individualsByGuid', 'familiesByGuid'})
        self.assertEqual(len(response_json['familiesByGuid']), 1)
        family_guid = next(iter(response_json['familiesByGuid'].keys()))
        self.assertEqual(response_json['familiesByGuid'][family_guid]['familyId'], 'PED073')
        self.assertEqual(response_json['familiesByGuid'][family_guid]['codedPhenotype'], 'Perinatal death')
        self.assertEqual(response_json['familiesByGuid'][family_guid]['mondoId'],  'MONDO:0100086')
        self.assertSetEqual(set(
            response_json['familiesByGuid'][family_guid]['individualGuids']),
            set(response_json['individualsByGuid'].keys())
        )
        self.assertSetEqual({i['familyGuid'] for i in response_json['individualsByGuid'].values()}, {family_guid})
        self.assertEqual(len(response_json['individualsByGuid']), 3)
        test_keys = {
            'affected', 'sex', 'notes', 'probandRelationship', 'primaryBiosample', 'analyteType', 'tissueAffectedStatus',
            'maternalId', 'paternalId'}
        indiv_1 = next(i for i in response_json['individualsByGuid'].values() if i['individualId'] == 'SCO_PED073B_GA0339_1')
        self.assertDictEqual({k: v for k, v in indiv_1.items() if k in test_keys}, {
            'affected': 'N', 'notes': 'probably dad', 'sex': 'M', 'maternalId': None, 'paternalId': None,
            'primaryBiosample': 'T', 'analyteType': 'B', 'tissueAffectedStatus': None,
            'probandRelationship': 'F',
        })
        indiv_2 = next(i for i in response_json['individualsByGuid'].values() if i['individualId'] == 'SCO_PED073C_GA0341_1')
        self.assertDictEqual({k: v for k, v in indiv_2.items() if k in test_keys}, {
            'affected': 'A', 'notes': 'half sib', 'sex': 'M', 'maternalId': None, 'paternalId': 'SCO_PED073B_GA0339_1',
            'primaryBiosample': 'BM', 'analyteType': 'R', 'tissueAffectedStatus': False,
            'probandRelationship': 'J',
        })
        indiv_3 = next(i for i in response_json['individualsByGuid'].values() if i['individualId'] == 'SCO_PED073C_GA0340_1')
        self.assertDictEqual({k: v for k, v in indiv_3.items() if k in test_keys}, {
            'affected': 'A', 'notes': None, 'sex': 'F', 'maternalId': None, 'paternalId': 'SCO_PED073B_GA0339_1',
             'primaryBiosample': 'BM', 'analyteType': 'D', 'tissueAffectedStatus': True, 'probandRelationship': 'S',
        })
        self.assertEqual(indiv_2['paternalGuid'], indiv_1['individualGuid'])
        self.assertEqual(indiv_3['paternalGuid'], indiv_1['individualGuid'])

    @mock.patch('seqr.views.utils.pedigree_info_utils.date')
    def test_individuals_datastat_table_handler(self, mock_date):
        mock_date.today.return_value = datetime.date(2020, 1, 1)

        receive_url = reverse(receive_individuals_table_handler, args=[PROJECT_GUID])
        self.check_manager_login(receive_url)

        data =  [['participant_guid', 'familyId', 'RELATIONSHIP', 'RELATIONSHIP_OTHER_DETAILS', 'WEBSITE', 'DESCRIPTION', 'CLINICAL_DIAGNOSES', 'CLINICAL_DIAGNOSES_DETAILS', 'GENETIC_DIAGNOSES', 'GENETIC_DIAGNOSES_DETAILS', 'FIND_OUT_DOCTOR_DETAILS', 'PATIENT_AGE', 'CONDITION_AGE', 'PATIENT_DECEASED', 'DECEASED_AGE', 'DECEASED_CAUSE', 'DECEASED_DNA', 'PATIENT_SEX', 'RACE', 'ETHNICITY', 'DOCTOR_TYPES', 'DOCTOR_TYPES_OTHER_DETAILS', 'TESTS', 'TESTS_MICROARRAY_YEAR', 'TESTS_MICROARRAY_LAB', 'TESTS_MICROARRAY_FAMILY', 'TESTS_MICROARRAY_FAMILY_OTHER_DETAILS',  'TESTS_WEXOME_YEAR', 'TESTS_WEXOME_LAB', 'TESTS_WEXOME_FAMILY', 'TESTS_WEXOME_FAMILY_OTHER_DETAILS', 'TESTS_WGENOME_YEAR', 'TESTS_WGENOME_LAB', 'TESTS_WGENOME_FAMILY', 'TESTS_WGENOME_FAMILY_OTHER_DETAILS', 'TESTS_OTHER_DETAILS', 'BIOPSY', 'BIOPSY_OTHER_DETAILS', 'OTHER_STUDIES', 'OTHER_STUDIES_DESCRIBE', 'EXPECT_RESULTS', 'MOTHER_SAME_CONDITION', 'MOTHER_CONDITION_AGE', 'MOTHER_RACE', 'MOTHER_ETHNICITY', 'MOTHER_CAN_PARTICIPATE', 'MOTHER_DECEASED', 'MOTHER_DECEASED_DNA', 'FATHER_SAME_CONDITION', 'FATHER_CONDITION_AGE', 'FATHER_RACE', 'FATHER_ETHNICITY', 'FATHER_CAN_PARTICIPATE', 'FATHER_DECEASED', 'FATHER_DECEASED_DNA', 'NO_SIBLINGS', 'SIBLING', 'NO_CHILDREN', 'CHILD', 'NO_RELATIVE_AFFECTED', 'RELATIVE', 'FAMILY_INFO'],
            ['1518231365', '123', 'OTHER', 'Grandchild', 'wwww.myblog.com', 'I have a really debilitating probably genetic condition. I\xe2ve seen many specialists.', 'YES', 'SMA\xe2s', 'YES', 'Dwarfism\xe2', 'Dr John Smith', '34', '21', 'YES', '33', 'heart attack', 'NO', 'MALE', 'WHITE,ASIAN,PACIFIC', 'NOT_HISPANIC', 'CLIN_GEN,NEURO,CARDIO,OTHER', 'Pediatrician', 'SINGLE_GENE,GENE_PANEL,WEXOME,WGENOME,OTHER', '', '', '', '', '2018', 'UDN\xe2s lab', 'PARENT,AUNT_UNCLE,NIECE_NEPHEW,OTHER', 'Grandmother',  '', '', '', 'Grandmother', 'Blood work', 'MUSCLE,SKIN,OTHER', 'Bone\xe2s', 'YES', 'Undiagnosed Diseases Network', 'NO', 'YES', '19', 'WHITE,ASIAN', 'NOT_HISPANIC', 'YES', '', '', 'NO', '', '', 'BLACK', 'PREFER_NOT_ANSWER', 'YES', 'NO', '', '[{"SIBLING_SEX":"FEMALE","SIBLING_AGE":"21","SIBLING_RACE":"WHITE","SIBLING_ETHNICITY":"NOT_HISPANIC","SIBLING_SAME_CONDITION":"YES","SIBLING_CONDITION_AGE":null,"SIBLING_CAN_PARTICIPATE":"NO"},{"SIBLING_SEX":"","SIBLING_AGE":"17","SIBLING_RACE": "WHITE","SIBLING_ETHNICITY":"NOT_HISPANIC","SIBLING_SAME_CONDITION":"","SIBLING_CONDITION_AGE":"","SIBLING_CAN_PARTICIPATE":"YES"}]', 'YES', '', 'NO', '[{"RELATIVE_SEX":"MALE","RELATIVE_AGE":"44","RELATIVE_RACE": "WHITE", "RELATIVE_ETHNICITY":"NOT_HISPANIC","RELATIVE_CONDITION_AGE":null,"RELATIVE_CAN_PARTICIPATE":null}]', 'patient\xe2s uncle (dads brother) died from Fahrs disease at 70'],
            ['b392fd78b440', '987', 'ADULT_CHILD', 'Grandchild', '', '', 'UNSURE', 'SMA', 'NO', 'Dwarfism', '', '47', '2', '', '33', 'heart attack', 'NO', 'PREFER_NOT_ANSWER', 'WHITE', 'UNKNOWN', '', 'Pediatrician', 'NOT_SURE,MICROARRAY,WEXOME', '', '', '', '', '2018', 'UDN', 'PARENT,AUNT_UNCLE,OTHER', 'Grandmother', '', '', '', 'Grandmother', 'Blood work', 'NONE', '', 'NO', 'Undiagnosed Diseases Network', 'NO', 'UNSURE', '19', '', 'UNKNOWN', 'NO', 'UNSURE', '', '', '', '', '', '', '', 'YES', 'YES', '[{"SIBLING_SEX":"FEMALE","SIBLING_AGE":"21","SIBLING_RACE":"WHITE","SIBLING_ETHNICITY":"NOT_HISPANIC","SIBLING_SAME_CONDITION":"YES","SIBLING_CONDITION_AGE":null,"SIBLING_CAN_PARTICIPATE":"NO"}]', 'NO', '[{"CHILD_SEX":"MALE","CHILD_AGE":"12","CHILD_RACE":"WHITE","CHILD_ETHNICITY":"NOT_HISPANIC","CHILD_SAME_CONDITION":"NO","CHILD_CONDITION_AGE":null,"CHILD_CAN_PARTICIPATE":"UNSURE"}]', 'YES', '', '']]
        response = self.client.post(receive_url, {'f': SimpleUploadedFile(
            'datstat.tsv', '\n'.join(['\t'.join(row) for row in data]).encode('utf-8')),
        })
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertDictEqual(response_json, {'uploadedFileId': mock.ANY, 'warnings': [], 'errors': [], 'info': [
            '2 families, 6 individuals parsed from datstat.tsv',
            '2 new families, 6 new individuals will be added to the project',
            '0 existing individuals will be updated',
        ]})

        url = reverse(save_individuals_table_handler, args=[PROJECT_GUID, response_json['uploadedFileId']])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {'individualsByGuid', 'familiesByGuid', 'familyNotesByGuid'})
        self.assertEqual(len(response_json['familiesByGuid']), 2)
        RGP_123_guid = next(f for f in response_json['familiesByGuid'].values() if f['familyId'] == 'RGP_123')['familyGuid']
        RGP_987_guid = next(f for f in response_json['familiesByGuid'].values() if f['familyId'] == 'RGP_987')['familyGuid']
        self.assertEqual(len(response_json['individualsByGuid']), 6)
        individuals = sorted(response_json['individualsByGuid'].values(), key=lambda i: i['individualId'])
        self.assertListEqual(
            [i['individualId'] for i in individuals],
            ['RGP_123_1', 'RGP_123_2', 'RGP_123_3', 'RGP_987_1', 'RGP_987_2', 'RGP_987_3'],
        )
        self.assertSetEqual({i['familyGuid'] for i in individuals[:3]}, {RGP_123_guid})
        self.assertSetEqual({i['familyGuid'] for i in individuals[3:]}, {RGP_987_guid})
        self.assertListEqual([i['sex'] for i in individuals], ['F', 'M', 'M', 'F', 'M', 'U'])
        self.assertListEqual([i['affected'] for i in individuals], ['N', 'N', 'A', 'N', 'N', 'A'])
        parents = individuals[:2] + individuals[3:5]
        no_parent_fields = [
            'maternalId', 'paternalId', 'maternalGuid', 'paternalGuid', 'maternalEthnicity', 'paternalEthnicity',
            'birthYear', 'deathYear', 'onsetAge', 'affectedRelatives',
        ]
        for no_parent_field in no_parent_fields:
            self.assertSetEqual({i[no_parent_field] for i in parents}, {None})

        self.assertDictEqual({k: v for k, v in individuals[2].items() if k in no_parent_fields}, {
            'maternalGuid': individuals[0]['individualGuid'], 'paternalGuid': individuals[1]['individualGuid'],
            'maternalId': 'RGP_123_1', 'paternalId': 'RGP_123_2', 'paternalEthnicity': ['Black'],
            'maternalEthnicity': ['White', 'Asian', 'Not Hispanic'], 'birthYear': 1986, 'deathYear': 2019,
            'onsetAge': 'A', 'affectedRelatives': True,
        })
        self.assertDictEqual({k: v for k, v in individuals[5].items() if k in no_parent_fields}, {
            'maternalGuid': individuals[3]['individualGuid'], 'paternalGuid': individuals[4]['individualGuid'],
            'maternalId': 'RGP_987_1', 'paternalId': 'RGP_987_2', 'maternalEthnicity': None, 'paternalEthnicity': None,
            'birthYear': 1973, 'deathYear': None, 'onsetAge': 'C', 'affectedRelatives': False,
        })

        self.assertEqual(len(response_json['familyNotesByGuid']), 2)
        note_1 = next(n['note'] for n in response_json['familyNotesByGuid'].values() if n['familyGuid'] == RGP_123_guid)
        self.assertEqual(note_1, """#### Clinical Information
* __Patient is my:__ Grandchild (male)
* __Current Age:__ Patient is deceased, age 33, due to heart attack, sample not available
* __Age of Onset:__ 21
* __Race/Ethnicity:__ White, Asian, Pacific; Not Hispanic
* __Case Description:__ I have a really debilitating probably genetic condition. Ive seen many specialists.
* __Clinical Diagnoses:__ Yes; SMAs
* __Genetic Diagnoses:__ Yes; Dwarfism
* __Website/Blog:__ Yes
* __Additional Information:__ patients uncle (dads brother) died from Fahrs disease at 70
#### Prior Testing
* __Referring Physician:__ Dr John Smith
* __Doctors Seen:__ Clinical geneticist, Neurologist, Cardiologist, Other: Pediatrician
* __Previous Testing:__ Yes;
* * Single gene testing
* * Gene panel testing
* * Whole exome sequencing. Year: 2018, Lab: UDNs lab, Relatives: Parent, Aunt or Uncle, Niece or Nephew, Other: Grandmother
* * Whole genome sequencing. Year: unspecified, Lab: unspecified, Relatives: None Specified
* * Other tests: Blood work
* __Biopsies Available:__ Muscle Biopsy, Skin Biopsy, Other Tissue Biopsy: Bones
* __Other Research Studies:__ Yes, Name of studies: Undiagnosed Diseases Network, Expecting results: No
#### Family Information
* __Mother:__ affected, onset age 19, available
* __Father:__ unaffected, unavailable, deceased, sample not available
* __Siblings:__ 
* * Sister, age 21, affected, unavailable
* * Sibling (unspecified sex), age 17, unspecified affected status, available
* __Children:__ None
* __Relatives:__ 
* * Male, age 44, affected, unspecified availability""")

        note_2 = next(n['note'] for n in response_json['familyNotesByGuid'].values() if n['familyGuid'] == RGP_987_guid)
        self.assertEqual(note_2, """#### Clinical Information
* __Patient is my:__ Adult Child (unspecified sex) - unable to provide consent
* __Current Age:__ 47
* __Age of Onset:__ 2
* __Race/Ethnicity:__ White; Unknown
* __Case Description:__ 
* __Clinical Diagnoses:__ Unsure
* __Genetic Diagnoses:__ No
* __Website/Blog:__ No
* __Additional Information:__ None specified
#### Prior Testing
* __Referring Physician:__ None
* __Doctors Seen:__ 
* __Previous Testing:__ Not sure
* __Biopsies Available:__ None
* __Other Research Studies:__ No
#### Family Information
* __Mother:__ unknown affected status, unavailable, unknown deceased status
* __Father:__ unknown affected status, unavailable, unspecified deceased status
* __Siblings:__ None
* __Children:__ 
* * Son, age 12, unaffected, unspecified availability
* __Relatives:__ None""")

    def _is_expected_individuals_metadata_upload(self, response, expected_families=False, has_non_hpo_update=False):
        unchanged_individuals = ['NA19679']
        if not has_non_hpo_update:
            unchanged_individuals.insert(0, 'NA19678')
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        expected_response = {
            'uploadedFileId': mock.ANY,
            'errors': [],
            'warnings': [
                "The following HPO terms were not found in seqr's HPO data and will not be added: HP:0004322 (NA19675_1); HP:0100258 (NA19679)",
                'Unable to find matching ids for 1 individuals. The following entries will not be updated: HG00731',
                f'No changes detected for {len(unchanged_individuals)} individuals. The following entries will not be updated: {", ".join(unchanged_individuals)}',
            ],
            'info': [f'{2 if has_non_hpo_update else 1} individuals will be updated'],
        }
        if expected_families:
            expected_response['warnings'].insert(1, 'The following invalid values for "assigned_analyst" will not be added: test_user_no_access@test.com (NA19679)')
        self.assertDictEqual(response_json, expected_response)

        # Save uploaded file
        url = reverse(save_individuals_metadata_table_handler, args=[PROJECT_GUID, response_json['uploadedFileId']])

        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        expected_keys = {'individualsByGuid'}
        if expected_families:
            expected_keys.add('familiesByGuid')
        self.assertSetEqual(set(response_json.keys()), expected_keys)
        updated_individuals = {'I000001_na19675'}
        if has_non_hpo_update:
            updated_individuals.add('I000002_na19678')
        self.assertSetEqual(set(response_json['individualsByGuid'].keys()), updated_individuals)
        self.assertSetEqual(set(response_json['individualsByGuid']['I000001_na19675'].keys()), INDIVIDUAL_FIELDS)
        self.assertListEqual(
            response_json['individualsByGuid']['I000001_na19675']['features'],
            [{'id': 'HP:0002017', 'category': 'HP:0025031', 'label': 'Nausea and vomiting'}]
        )
        self.assertListEqual(
            response_json['individualsByGuid']['I000001_na19675']['absentFeatures'],
            [{'id': 'HP:0012469', 'category': 'HP:0025031', 'label': 'Infantile spasms'}]
        )
        self.assertEqual(response_json['individualsByGuid']['I000001_na19675']['sex'], 'XXY')
        self.assertEqual(response_json['individualsByGuid']['I000001_na19675']['birthYear'], 2000)
        self.assertTrue(response_json['individualsByGuid']['I000001_na19675']['affectedRelatives'])
        self.assertEqual(response_json['individualsByGuid']['I000001_na19675']['onsetAge'], 'J')
        self.assertListEqual(response_json['individualsByGuid']['I000001_na19675']['expectedInheritance'], ['D', 'S'])
        self.assertListEqual(
            response_json['individualsByGuid']['I000001_na19675']['maternalEthnicity'], ['Finnish', 'Irish'])
        self.assertListEqual(
            response_json['individualsByGuid']['I000001_na19675']['candidateGenes'],
            [{'gene': 'IKBKAP', 'comments': 'multiple panels, no confirm'}, {'gene': 'EHBP1L1'}])

        if has_non_hpo_update:
            self.assertIsNone(response_json['individualsByGuid']['I000002_na19678']['features'])
            self.assertFalse(response_json['individualsByGuid']['I000002_na19678']['affectedRelatives'])

        if expected_families:
            self.assertListEqual(list(response_json['familiesByGuid'].keys()), ['F000001_1'])
            self.assertDictEqual(
                response_json['familiesByGuid']['F000001_1']['assignedAnalyst'],
                {'email': 'test_user_collaborator@test.com', 'fullName': 'Test Collaborator User'}
            )

    def test_individuals_metadata_table_handler(self):
        url = reverse(receive_individuals_metadata_handler, args=['R0001_1kg'])
        self.check_collaborator_login(url)

        # Send invalid requests
        header = 'family_id,indiv_id,hpo_term_yes,hpo_term_no'
        f = SimpleUploadedFile('updates.csv', header.encode('utf-8'))
        response = self.client.post(url, data={'f': f})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {'errors': ['Invalid header, missing individual id column'], 'warnings': []})

        header = 'family_id,individual_id,hpo_term_present,hpo_term_absent,sex,birth year,other affected relatives,onset age,expected inheritance,maternal ancestry,candidate genes,assigned analyst'
        rows = [
            '1,NA19678,,,,,no,infant,recessive,,,not_an_email',
            '1,NA19679,HP:0100258 (Preaxial polydactyly),,,,,,,,,test_user_no_access@test.com',
            '1,HG00731,HP:0002017,HP:0012469 (Infantile spasms);HP:0011675 (Arrhythmia);HP:0011675 (Arrhythmia),,,,,,,,,',
        ]
        f = SimpleUploadedFile('updates.csv', "{}\n{}".format(header, '\n'.join(rows)).encode('utf-8'))
        response = self.client.post(url, data={'f': f})
        self.assertEqual(response.status_code, 400)
        self.assertDictEqual(response.json(), {
            'errors': [
                'Unable to find individuals to update for any of the 3 parsed individuals. No matching ids found for 1 individuals. No changes detected for 2 individuals.'
            ],
            'warnings': [
                "The following HPO terms were not found in seqr's HPO data and will not be added: HP:0100258 (NA19679)",
                'The following invalid values for "affected_relatives" will not be added: no (NA19678)',
                'The following invalid values for "onset_age" will not be added: infant (NA19678)',
                'The following invalid values for "expected_inheritance" will not be added: recessive (NA19678)',
                'The following invalid values for "assigned_analyst" will not be added: not_an_email (NA19678); test_user_no_access@test.com (NA19679)',
                'Unable to find matching ids for 1 individuals. The following entries will not be updated: HG00731',
                'No changes detected for 2 individuals. The following entries will not be updated: NA19678, NA19679',
            ]})

        # send valid request
        rows[0] = '1,NA19678,,,,,false,,,,,'
        rows.append('1,NA19675_1,HP:0002017,"HP:0012469 (Infantile spasms);HP:0004322 (Short stature, severe)",F,2000,True,Juvenile onset,"Autosomal dominant inheritance, Sporadic","Finnish, Irish","IKBKAP -- (multiple panels, no confirm), EHBP1L1",test_user_collaborator@test.com')
        f = SimpleUploadedFile('updates.csv', "{}\n{}".format(header, '\n'.join(rows)).encode('utf-8'))
        response = self.client.post(url, data={'f': f})
        self._is_expected_individuals_metadata_upload(response, expected_families=True, has_non_hpo_update=True)

    def test_individuals_metadata_hpo_term_number_table_handler(self):
        url = reverse(receive_individuals_metadata_handler, args=['R0001_1kg'])
        self.check_collaborator_login(url)

        header = 'family_id,individual_id,affected,hpo_number,hpo_number,sex,birth,other affected relatives,onset,expected inheritance,maternal ancestry,candidate genes'
        rows = [
            '1,NA19675_1,yes,HP:0002017,,F,2000,true,Juvenile onset,"Autosomal dominant inheritance, Sporadic","Finnish, Irish","IKBKAP -- (multiple panels, no confirm), EHBP1L1"',
            '1,NA19675_1,no,HP:0012469,,,,,,,,',
            '1,NA19675_1,no,HP:0012469,,,,,,,,',
            '1,NA19675_1,no,,HP:0004322,,,,,,,',
            '1,NA19678,,,,,,,,,,',
            '1,NA19679,yes,HP:0100258,,,,,,,,',
            '1,NA19679,yes,HP:0100258,,,,,,,,',
            '1,HG00731,yes,HP:0002017,,,,,,,,',
            '1,HG00731,no,HP:0012469,HP:0011675,,,,,,,',
        ]
        f = SimpleUploadedFile('updates.csv', "{}\n{}".format(header, '\n'.join(rows)).encode('utf-8'))
        response = self.client.post(url, data={'f': f})
        self._is_expected_individuals_metadata_upload(response)

    def _set_metadata_file_iter(self, mock_subprocess, genetic_findings_table):
        mock_subprocess.return_value.stdout.__iter__.side_effect = [
            iter(['\t'.join(row).encode() for row in file]) for file in [
                EXPERIMENT_TABLE, EXPERIMENT_LOOKUP_TABLE, LOAD_PARTICIPANT_TABLE, PHENOTYPE_TABLE,
                genetic_findings_table,
            ]
        ]

    @mock.patch('seqr.views.utils.permissions_utils.PM_USER_GROUP', 'project-managers')
    @mock.patch('seqr.utils.file_utils.subprocess.Popen')
    def test_import_gregor_metadata(self, mock_subprocess):
        genetic_findings_table = deepcopy(GENETIC_FINDINGS_TABLE)
        genetic_findings_table[2] = genetic_findings_table[2][:11] + genetic_findings_table[4][11:14] + \
                                    genetic_findings_table[2][14:]
        genetic_findings_table.append([
            'Broad_NA20889_1_249045487', 'Broad_NA20889', '', 'SNV/INDEL', 'GRCh37', '1', '249045487', 'A', 'G', '',
            'OR4G11P', '', '', '', 'Heterozygous', '', 'unknown', 'Broad_NA20889_1_248367227', '', 'Candidate',
            'IRIDA syndrome', 'MONDO:0008788', 'Autosomal dominant', 'Full', '', '', 'SR-ES', '', '', '', '', '', '', '',
        ])
        self._set_metadata_file_iter(mock_subprocess, genetic_findings_table)

        url = reverse(import_gregor_metadata, args=[PM_REQUIRED_PROJECT_GUID])
        self.check_pm_login(url)

        body = {
            'workspaceNamespace': 'my-seqr-billing', 'workspaceName': 'anvil-1kg project nåme with uniçøde',
            'sampleType': 'exome',
        }
        response = self.client.post(url, content_type='application/json', data=json.dumps(body))
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {
            'importStats', 'projectsByGuid', 'familiesByGuid', 'individualsByGuid', 'familyTagTypeCounts',
        })
        warnings = [
            'Broad_HG00733 is the mother of VCGS_FAM203_621_D2 but is not included',
            'Skipped the following unrecognized HPO terms: HP:0001509',
        ]
        self.assertDictEqual(response_json['importStats'], {'gregorMetadata': {
            'warnings': warnings, 'info': [
                'Imported 4 individuals',
                'Created 1 new families, 3 new individuals',
                'Updated 1 existing families, 1 existing individuals',
                'Skipped 0 unchanged individuals',
                'Loaded 4 new and 0 updated findings tags',
            ],
        }})

        self.assertDictEqual(response_json['projectsByGuid'], {
            PM_REQUIRED_PROJECT_GUID: {'variantTagTypes': mock.ANY, 'variantFunctionalTagTypes': mock.ANY},
        })
        self.assertDictEqual(response_json['projectsByGuid'][PM_REQUIRED_PROJECT_GUID]['variantTagTypes'][1], {
            'variantTagTypeGuid': 'VTT_gregor_finding',
            'name': 'GREGoR Finding',
            'category': '',
            'description': '',
            'metadataTitle': None,
            'color': '#c25fc4',
            'order': 0.5,
            'numTags': 5,
        })

        self.assertEqual(len(response_json['familiesByGuid']), 2)
        self.assertIn('F000012_12', response_json['familiesByGuid'])
        new_family_guid = next(k for k in response_json['familiesByGuid'] if k != 'F000012_12')
        self.assertEqual(response_json['familiesByGuid'][new_family_guid]['familyId'], 'Broad_2')
        self.assertEqual(response_json['familiesByGuid'][new_family_guid]['codedPhenotype'], 'microcephaly; seizures')

        self.assertDictEqual(response_json['familyTagTypeCounts'], {
            'F000012_12': {'GREGoR Finding': 3, 'MME Submission': 2, 'Tier 1 - Novel gene and phenotype': 1},
            new_family_guid: {'GREGoR Finding': 2},
        })

        self.assertEqual(len(response_json['individualsByGuid']), 4)
        self.assertIn('I000016_na20888', response_json['individualsByGuid'])
        created_individual_guid = next(
            k for k, v in response_json['individualsByGuid'].items() if v['individualId'] == 'Broad_NA20889')
        new_family_individual_guid = next(
            k for k, v in response_json['individualsByGuid'].items() if v['individualId'] == 'VCGS_FAM203_621_D2')

        individual_db_data = Individual.objects.filter(
            guid__in=response_json['individualsByGuid']).order_by('individual_id').values(
            'individual_id', 'display_name', 'family__guid', 'affected', 'sex', 'proband_relationship', 'population',
            'mother__individual_id', 'father__individual_id', 'features', 'absent_features', 'case_review_status',
        )
        self.assertDictEqual(individual_db_data[0], {
            'individual_id': 'Broad_HG00732',
            'display_name': '',
            'family__guid': new_family_guid,
            'affected': 'N',
            'sex': 'M',
            'proband_relationship': 'F',
            'mother__individual_id': None,
            'father__individual_id': None,
            'population': 'NFE',
            'features': [],
            'absent_features': [],
            'case_review_status': 'I',
        })
        self.assertDictEqual(individual_db_data[1], {
            'individual_id': 'Broad_NA20889',
            'display_name': '',
            'family__guid': 'F000012_12',
            'affected': 'A',
            'sex': 'F',
            'proband_relationship': 'S',
            'mother__individual_id': None,
            'father__individual_id': None,
            'population': 'OTH',
            'features': [{'id': 'HP:0011675'}],
            'absent_features': [],
            'case_review_status': 'I',
        })
        self.assertDictEqual(individual_db_data[2], {
            'individual_id': 'NA20888',
            'display_name': 'Broad_NA20888',
            'family__guid': 'F000012_12',
            'affected': 'A',
            'sex': 'M',
            'proband_relationship': '',
            'mother__individual_id': None,
            'father__individual_id': None,
            'population': 'SAS',
            'features': [],
            'absent_features': [],
            'case_review_status': 'G',
        })
        self.assertDictEqual(individual_db_data[3], {
            'individual_id': 'VCGS_FAM203_621_D2',
            'display_name': 'Broad_HG00731',
            'family__guid': new_family_guid,
            'affected': 'A',
            'sex': 'F',
            'proband_relationship': 'S',
            'mother__individual_id': None,
            'father__individual_id': 'Broad_HG00732',
            'population': 'AMR',
            'features': [{'id': 'HP:0011675'}],
            'absent_features': [{'id': 'HP:0002017'}],
            'case_review_status': 'I',
        })

        saved_variants = SavedVariant.objects.filter(
            varianttag__variant_tag_type__name='GREGoR Finding'
        ).order_by('family_id', 'variant_id').distinct().values(
            'guid', 'variant_id', 'xpos', 'family__guid', 'saved_variant_json__genomeVersion',
            'saved_variant_json__transcripts', 'saved_variant_json__genotypes', 'saved_variant_json__mainTranscriptId',
            'saved_variant_json__hgvsc',
        )
        self.assertEqual(len(saved_variants), 4)
        self.assertDictEqual(saved_variants[0], {
            'guid': 'SV0000006_1248367227_r0003_tes',
            'variant_id': '1-248367227-TC-T',
            'xpos': 1248367227,
            'family__guid': 'F000012_12',
            'saved_variant_json__genomeVersion': '37',
            'saved_variant_json__transcripts': mock.ANY,
            'saved_variant_json__genotypes': mock.ANY,
            'saved_variant_json__mainTranscriptId': 'ENST00000505820',
            'saved_variant_json__hgvsc': None,
        })
        self.assertEqual(len(saved_variants[0]['saved_variant_json__transcripts']), 2)
        self.assertEqual(len(saved_variants[0]['saved_variant_json__genotypes']), 2)
        self.assertDictEqual(saved_variants[1], {
            'guid': mock.ANY,
            'variant_id': '1-249045487-A-G',
            'xpos': 1249045487,
            'family__guid': 'F000012_12',
            'saved_variant_json__genomeVersion': '37',
            'saved_variant_json__transcripts': {
                'ENSG00000240361': [{'hgvsc': None, 'hgvsp': None, 'transcriptId': None}],
            },
            'saved_variant_json__genotypes': {created_individual_guid: {'numAlt': 1}},
            'saved_variant_json__mainTranscriptId': None,
            'saved_variant_json__hgvsc': None,
        })
        new_family_genotypes = {new_family_individual_guid: {'numAlt': 2}}
        self.assertDictEqual(saved_variants[2], {
            'guid': mock.ANY,
            'variant_id': '1-248367227-TC-T',
            'xpos': 1248367227,
            'family__guid': new_family_guid,
            'saved_variant_json__genomeVersion': '37',
            'saved_variant_json__transcripts': {
                'ENSG00000135953': [{'hgvsc': 'c.3955G>A', 'hgvsp': 'c.1586-17C>G', 'transcriptId': 'ENST00000505820'}]
            },
            'saved_variant_json__genotypes': new_family_genotypes,
            'saved_variant_json__mainTranscriptId': 'ENST00000505820',
            'saved_variant_json__hgvsc': None,
        })

        variant_tags = VariantTag.objects.filter(variant_tag_type__name='GREGoR Finding')
        existing_variant_tags = variant_tags.filter(saved_variants__guid='SV0000006_1248367227_r0003_tes')
        new_variant_tags = variant_tags.filter(saved_variants__guid=saved_variants[1]['guid'])
        comp_het_tags = set(existing_variant_tags).intersection(new_variant_tags)
        self.assertEqual(len(comp_het_tags), 1)
        comp_het_tag = comp_het_tags.pop()
        self.assertIsNone(comp_het_tag.metadata)
        self.assertDictEqual(json.loads(next(t for t in existing_variant_tags if t != comp_het_tag).metadata), {
            'gene_known_for_phenotype': 'Candidate',
            'condition_id': 'OMIM:616126',
            'known_condition_name': 'Immunodeficiency 38',
            'condition_inheritance': 'Autosomal recessive',
        })
        self.assertDictEqual(json.loads(next(t for t in new_variant_tags if t != comp_het_tag).metadata), {
            'gene_known_for_phenotype': 'Candidate',
            'condition_id': 'MONDO:0008788',
            'known_condition_name': 'IRIDA syndrome',
            'condition_inheritance': 'Autosomal dominant',
        })

        new_family_tag = variant_tags.get(saved_variants__guid=saved_variants[2]['guid'])
        self.assertDictEqual(
            json.loads(new_family_tag.metadata), {'gene_known_for_phenotype': 'Known', 'condition_id': 'MONDO:0044970'},
        )

        mock_subprocess.assert_has_calls([
            mock.call('gsutil cat gs://test_bucket/data_tables/experiment_dna_short_read.tsv', stdout=-1, stderr=-2, shell=True),  # nosec
            mock.call().stdout.__iter__(),
            mock.call('gsutil cat gs://test_bucket/data_tables/experiment.tsv', stdout=-1, stderr=-2, shell=True),  # nosec
            mock.call().stdout.__iter__(),
            mock.call('gsutil cat gs://test_bucket/data_tables/participant.tsv', stdout=-1, stderr=-2, shell=True), # nosec
            mock.call().stdout.__iter__(),
            mock.call('gsutil cat gs://test_bucket/data_tables/phenotype.tsv', stdout=-1, stderr=-2, shell=True),  # nosec
            mock.call().stdout.__iter__(),
            mock.call('gsutil cat gs://test_bucket/data_tables/genetic_findings.tsv', stdout=-1, stderr=-2, shell=True),  # nosec
            mock.call().stdout.__iter__(),
        ])

        # Test behavior on reload
        SavedVariant.objects.get(guid=saved_variants[2]['guid']).delete()
        genetic_findings_table[2][10] = 'PPX123'
        self._set_metadata_file_iter(mock_subprocess, genetic_findings_table)
        response = self.client.post(url, content_type='application/json', data=json.dumps(body))
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        warnings.append('The following unknown genes were omitted in the findings tags: PPX123')
        self.assertDictEqual(response_json['importStats'], {'gregorMetadata': {
            'warnings': warnings, 'info': [
                'Imported 4 individuals',
                'Created 0 new families, 0 new individuals',
                'Updated 0 existing families, 0 existing individuals',
                'Skipped 4 unchanged individuals',
                'Loaded 1 new and 3 updated findings tags',
            ],
        }})
        self.assertDictEqual(response_json['individualsByGuid'], {})

        no_gene_saved_variant_json = SavedVariant.objects.get(family__guid=new_family_guid, variant_id='1-248367227-TC-T').saved_variant_json
        self.assertDictEqual(no_gene_saved_variant_json['transcripts'], {})
        self.assertDictEqual(no_gene_saved_variant_json['genotypes'], new_family_genotypes)
        self.assertNotIn('mainTranscriptId', no_gene_saved_variant_json)
        self.assertNotIn('hgvsc', no_gene_saved_variant_json)

    def test_get_hpo_terms(self):
        url = reverse(get_hpo_terms, args=['HP:0011458'])
        self.check_require_login(url)

        response = self.client.get(url, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertDictEqual(response.json(), {
            'HP:0011458': {
                'HP:0002017': {'id': 'HP:0002017', 'category': 'HP:0025031', 'label': 'Nausea and vomiting'},
                'HP:0001252': {'id': 'HP:0001252', 'category': 'HP:0025031', 'label': 'Muscular hypotonia'},
            }
        })

        url = reverse(get_hpo_terms, args=['HP:0002017'])
        response = self.client.get(url, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertDictEqual(response.json(), {
            'HP:0002017': {}
        })

    def test_get_individual_rna_seq_data(self):
        url = reverse(get_individual_rna_seq_data, args=[INDIVIDUAL_GUID])
        self.check_collaborator_login(url)

        response = self.client.get(url, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        response_json = response.json()
        self.assertSetEqual(set(response_json.keys()), {'rnaSeqData', 'genesById'})
        self.assertDictEqual(response_json['rnaSeqData'], {
            INDIVIDUAL_GUID: {'outliers': {
                'ENSG00000135953': [{
                    'geneId': 'ENSG00000135953', 'zScore': 7.31, 'pValue': 0.00000000000948, 'pAdjust': 0.00000000781,
                    'isSignificant': True,
                    'tissueType': 'M',
                }],
                'ENSG00000240361': [{
                    'geneId': 'ENSG00000240361', 'zScore': -4.08, 'pValue': 5.88, 'pAdjust': 0.09, 'isSignificant': False,
                    'tissueType': 'M',
                }],
                'ENSG00000268903': [{
                    'geneId': 'ENSG00000268903', 'zScore': 7.08, 'pValue':0.000000000588, 'pAdjust': 0.00000000139,
                    'isSignificant': True,
                    'tissueType': 'M',
                }],
            },
            'spliceOutliers': {
                'ENSG00000268903': mock.ANY,
            },
        }})
        outliers_by_pos = {outlier['start']: outlier for outlier in
                           response_json['rnaSeqData'][INDIVIDUAL_GUID]['spliceOutliers']['ENSG00000268903']}
        self.assertDictEqual(
            {
                'chrom': '7', 'counts': 1297, 'end': 132886973, 'geneId': 'ENSG00000268903', 'isSignificant': True,
                'meanCounts': 0.85,  'meanTotalCounts': 0.85, 'pAdjust': 3.08e-56,
                'pValue': 1.08e-56, 'rareDiseaseSamplesTotal': 20, 'rareDiseaseSamplesWithThisJunction': 1,
                'totalCounts': 1297, 'start': 132885746, 'strand': '*', 'type': 'psi5', 'deltaIntronJaccardIndex': 12.34,
                'tissueType': 'F',
            },
            outliers_by_pos[132885746]
        )
        self.assertSetEqual(set(response_json['genesById'].keys()), {'ENSG00000135953', 'ENSG00000268903'})
        self.assertSetEqual(set(response_json['genesById']['ENSG00000135953'].keys()), GENE_FIELDS)

    def test_get_individual_rna_seq_data_is_significant(self):
        url = reverse(get_individual_rna_seq_data, args=[INDIVIDUAL_GUID])
        self.check_collaborator_login(url)

        response = self.client.get(url, content_type='application/json')
        self.assertEqual(response.status_code, 200)
        response_rnaseq_data = response.json()['rnaSeqData'][INDIVIDUAL_GUID]
        self.assertTrue(response_rnaseq_data['outliers']['ENSG00000135953'][0]['isSignificant'])
        significant_outliers = [outlier for outlier in response_rnaseq_data['outliers'].values() if outlier[0]['isSignificant']]
        self.assertEqual(2, len(significant_outliers))
        self.assertListEqual(
            sorted([{field: outlier[field] for field in ['start', 'end', 'pAdjust', 'deltaIntronJaccardIndex', 'tissueType', 'isSignificant']}
                    for outlier in response_rnaseq_data['spliceOutliers']['ENSG00000268903']], key=lambda r: r['start']),
            [{'start': 1001, 'end': 2001, 'pAdjust': 0.3, 'deltaIntronJaccardIndex': 12.34, 'tissueType': 'F', 'isSignificant': False},
             {'start': 3000, 'end': 4000, 'pAdjust': 0.0003, 'deltaIntronJaccardIndex': -12.34, 'tissueType': 'F', 'isSignificant': True},
             {'start': 5000, 'end': 6000, 'pAdjust': 0.0003, 'deltaIntronJaccardIndex': 0.05, 'tissueType': 'F', 'isSignificant': False},
             {'start': 7000, 'end': 8000, 'pAdjust': 0.003, 'deltaIntronJaccardIndex': 12.34, 'tissueType': 'M', 'isSignificant': True},
             {'start': 9000, 'end': 9100, 'pAdjust': 0.1, 'deltaIntronJaccardIndex': -0.01, 'tissueType': 'M', 'isSignificant': False},
             {'start': 132885746, 'end': 132886973, 'pAdjust': 3.08e-56, 'deltaIntronJaccardIndex': 12.34, 'tissueType': 'F', 'isSignificant': True}],
        )


class LocalIndividualAPITest(AuthenticationTestCase, IndividualAPITest):
    fixtures = ['users', '1kg_project', 'reference_data']
    HAS_EXTERNAL_PROJECT_ACCESS = False

    def setUp(self):
        patcher = mock.patch('seqr.utils.file_utils.subprocess.Popen')
        _mock_subprocess = patcher.start()
        _mock_subprocess.side_effect = Exception('Calling gs from local')
        self.addCleanup(patcher.stop)

        super().setUp()

    def test_import_gregor_metadata(self, *args):
        # Importing gregor metadata does not work in local environment
        pass


class AnvilIndividualAPITest(AnvilAuthenticationTestCase, IndividualAPITest):
    fixtures = ['users', 'social_auth', '1kg_project', 'reference_data']
    HAS_EXTERNAL_PROJECT_ACCESS = True

    def setUp(self):
        patcher = mock.patch('seqr.utils.file_utils.subprocess.Popen')
        _mock_subprocess = patcher.start()
        self.addCleanup(patcher.stop)

        self.mock_subprocess = mock.MagicMock()
        self.mock_subprocess.wait.return_value = 0
        self.mock_subprocess.stdout.__iter__.return_value = []
        self.gs_files = {}
        _mock_subprocess.side_effect = self._mock_subprocess

        super().setUp()

    def _mock_subprocess(self, command, **kwargs):
        command_args = re.match(
            r'gsutil (?P<cmd>cat|mv)(?P<local_path> \S+)? gs://seqr-scratch-temp/(?P<gs_path>\S+)', command,
        ).groupdict()
        file_name = command_args['gs_path']
        if command_args['cmd'] == 'mv':
            src_path = command_args['local_path'].strip()
            self.assertEqual(src_path.split('/')[-1], file_name)
            with gzip.open(src_path) as f:
                self.gs_files[file_name] = f.readlines()
        else:
            self.mock_subprocess.stdout.__iter__.return_value = self.gs_files[file_name]
        return self.mock_subprocess

    def _assert_expected_delete_individuals(self, response, mock_pm_group):
        self.assertEqual(response.status_code, 400)
        self.assertListEqual(response.json()['errors'], ['Unable to delete individuals with active search sample: NA19678'])
